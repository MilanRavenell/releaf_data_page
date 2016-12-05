#!/usr/bin/python

from sklearn import datasets, cluster, preprocessing, metrics
from matplotlib import pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import openpyxl
import pandas as pd
from sklearn.feature_extraction import DictVectorizer
import sys

def is_number(s):
    
    if s != "Unknown" and s != None:
    	return True
    else:
    	return False

def column(matrix, i):
    return [row[i] for row in matrix]

def zip(array, column):
	result = array

	if len(result) == 0:
		return column
	else: 
		for i in range(len(result)):
			for j in range(len(column[i])):
				result[i].append(column[i][j])
		return result

def order(array, clusters):
	occurences = []
	lookup = []
	
	for i in range(clusters):
		occurences.append((array.count(i), i))
	

	sorted_occurences = sorted(occurences, key=lambda tup: tup[0], reverse = True)

	for i in range(len(array)):
		array[i] = [x for x, y in enumerate(sorted_occurences) if y[1] == array[i]][0]

def main(argv):

	X = []
	X_cat = []
	y = []

	wb = openpyxl.load_workbook('uploads/' + argv[0])
	sheet = wb.get_sheet_by_name(argv[1])

	for i in range(2, 1000):

		city = sheet.cell(row = i, column = 4).value
		perp = sheet.cell(row = i, column = 5).value
		fatalities = sheet.cell(row = i, column = 11).value
		injured = sheet.cell(row = i, column = 12).value
		target = sheet.cell(row = i, column = 13).value
		attack = sheet.cell(row = i, column = 17).value
		weapon = sheet.cell(row = i, column = 20).value

		if is_number(argv[2]) and is_number(argv[3]): 

			X.append([int(argv[3]), int(argv[2])])
			X_cat.append([argv[4]])
			y.append(argv[5])

			le = preprocessing.LabelEncoder()
			y0 = le.fit_transform(y)

	clusters = max(y0) + 1
	print "\nLegend\n"
	for i in range(clusters):
		print i, le.inverse_transform(i)

	if X_cat:
		for i in range(len(X_cat[0])):

			le1 = preprocessing.LabelEncoder()
			X_cat0 = le.fit_transform(column(X_cat, i))
			
			lb = preprocessing.LabelBinarizer()
			X_cat0 = lb.fit_transform(X_cat0)

			if not len(X) == 0:
				X = zip(X, X_cat0)
			else:
				X = X_cat0

	k_means = cluster.KMeans(n_clusters = clusters)
	k_means.fit(X)
	result = k_means.labels_.tolist()

	y1 = y0.tolist()
	order(result, clusters)
	order(y1, clusters)

	score = metrics.adjusted_rand_score(result, y1)  
	print "\nRand Score: ", score

	fig = plt.figure()
	ax = fig.add_subplot(111, projection='3d')
	x_d = column(X, 0)
	y_d = column(X,1)
	z_d = lb.inverse_transform(X_cat0)

	ax.scatter(x_d, y_d, z_d, c= y1)

	plt.xlabel('fatalities')
	plt.ylabel('injuries')
	ax.set_zlabel('weapon')

	plt.show()
	plot(result, y1)

def plot(X, y):
	plt.subplot(211)
	plt.scatter(y, range(0, len(y)))
	plt.title('target')
	
	plt.subplot(212)
	plt.scatter(X, range(0, len(y)))
	plt.title('cluster')

	plt.show()

# sharif()
if __name__ == "__main__":
	main(sys.argv[1:])




